"""Aggregate volume provenance and exchange identity evidence."""

from __future__ import annotations

import json
from collections import Counter
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from alphapilot.data_provenance.volume_semantics_verifier import (
    verify_volume_semantics,
)
from alphapilot.evolution.registry.hashing import sha256_file, stable_hash


_VOLUME_COLUMN_PRIORITY = (
    "volume_quote_currency",
    "volCcyQuote",
    "quoteVolume",
    "volume_base_currency",
    "volCcy",
    "baseVolume",
    "volume_base_or_contracts",
    "vol",
    "volume",
)


def _manifest_rows(payload: Mapping[str, Any]) -> list[dict[str, Any]]:
    for key in ("rows", "datasets", "assets", "partitions"):
        values = payload.get(key)
        if isinstance(values, list):
            return [dict(value) for value in values if isinstance(value, Mapping)]
    return []


def _raw_file(raw_root: Path, instrument: str, timeframe: str) -> Path | None:
    symbol_dir = instrument.replace("-", "_")
    timeframe_dir = f"swap_candles_{timeframe.upper()}"
    directory = raw_root / timeframe_dir / symbol_dir
    preferred = directory / f"{symbol_dir}_ALL.xlsx"
    if preferred.is_file():
        return preferred
    combined = sorted(directory.glob("*_ALL.xlsx")) if directory.is_dir() else []
    if combined:
        return combined[0]
    matches = sorted(directory.glob("*.xlsx")) if directory.is_dir() else []
    return matches[0] if matches else None


def _canonical_file(
    canonical_root: Path, instrument: str, timeframe: str
) -> Path | None:
    directory = canonical_root / instrument / timeframe.lower()
    matches = sorted(directory.glob("*.parquet")) if directory.is_dir() else []
    return matches[0] if matches else None


def _xlsx_header(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(value) for value in row if value is not None]
    finally:
        workbook.close()


def _declared_unit(column: str) -> str:
    if column in {"volume_quote_currency", "volCcyQuote", "quoteVolume"}:
        return "quote_asset"
    if column in {"volume_base_currency", "volCcy", "baseVolume"}:
        return "base_asset"
    if column == "volume_base_or_contracts":
        return "unknown_base_or_contracts"
    return "unknown"


def discover_volume_provenance_records(
    *,
    manifest_path: Path,
    raw_root: Path,
    canonical_root: Path,
    instruments: Sequence[str],
    timeframes: Sequence[str],
) -> list[dict[str, Any]]:
    """Bind source headers, reader mapping, hashes, and manifest metadata."""

    manifest_path = Path(manifest_path).resolve()
    raw_root = Path(raw_root).resolve()
    canonical_root = Path(canonical_root).resolve()
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    rows = _manifest_rows(payload)
    index = {
        (
            str(
                row.get("instrumentId")
                or row.get("instrument")
                or ((row.get("symbols") or [""])[0])
            ),
            str(row.get("timeframe") or "").lower(),
        ): row
        for row in rows
        if str(row.get("datasetType") or row.get("kind") or "ohlcv").lower()
        == "ohlcv"
    }
    records: list[dict[str, Any]] = []
    for instrument in instruments:
        for timeframe in timeframes:
            normalized_timeframe = str(timeframe).lower()
            raw_path = _raw_file(raw_root, instrument, normalized_timeframe)
            canonical_path = _canonical_file(
                canonical_root, instrument, normalized_timeframe
            )
            manifest_row = index.get((instrument, normalized_timeframe), {})
            if raw_path is None:
                records.append(
                    {
                        "datasetId": f"{instrument}_{normalized_timeframe}",
                        "instrumentId": instrument,
                        "timeframe": normalized_timeframe,
                        "sourceExchange": str(
                            manifest_row.get("exchange") or "unknown"
                        ),
                        "rawColumnNames": [],
                        "selectedVolumeColumn": "",
                        "selectedVolumeColumnIndex": None,
                        "declaredVolumeUnit": "unknown",
                        "sourceFileHash": "",
                        "rawPath": None,
                        "canonicalPath": str(canonical_path.resolve())
                        if canonical_path
                        else None,
                        "canonicalReaderMapping": None,
                        "rowCount": int(manifest_row.get("rowCount") or 0),
                        "start": manifest_row.get("start")
                        or manifest_row.get("startTime"),
                        "end": manifest_row.get("end")
                        or manifest_row.get("endTime"),
                        "evidenceRefs": ["manifest"] if manifest_row else [],
                        "limitations": ["raw source file not found"],
                    }
                )
                continue
            columns = _xlsx_header(raw_path)
            selected = next(
                (column for column in _VOLUME_COLUMN_PRIORITY if column in columns),
                "",
            )
            evidence = ["raw_multicolumn_file"]
            if manifest_row:
                evidence.append("manifest")
            mapping = None
            if selected and canonical_path is not None:
                mapping = {
                    "canonicalVolumeField": "volume",
                    "rawVolumeField": selected,
                }
                evidence.append("canonical_reader_mapping")
            records.append(
                {
                    "datasetId": f"{instrument}_{normalized_timeframe}",
                    "instrumentId": instrument,
                    "timeframe": normalized_timeframe,
                    "sourceExchange": str(
                        manifest_row.get("exchange") or "unknown"
                    ),
                    "marketType": str(
                        manifest_row.get("marketType") or "unknown"
                    ),
                    "rawColumnNames": columns,
                    "selectedVolumeColumn": selected,
                    "selectedVolumeColumnIndex": columns.index(selected)
                    if selected
                    else None,
                    "declaredVolumeUnit": _declared_unit(selected),
                    "sourceFileHash": sha256_file(raw_path),
                    "contentHash": str(
                        manifest_row.get("contentHash")
                        or manifest_row.get("actualContentHash")
                        or ""
                    ),
                    "rawPath": str(raw_path.resolve()),
                    "canonicalPath": str(canonical_path.resolve())
                    if canonical_path
                    else None,
                    "canonicalReaderMapping": mapping,
                    "rowCount": int(manifest_row.get("rowCount") or 0),
                    "start": manifest_row.get("start")
                    or manifest_row.get("startTime"),
                    "end": manifest_row.get("end")
                    or manifest_row.get("endTime"),
                    "availableAtRule": "candle_close_timestamp",
                    "evidenceRefs": evidence,
                    "limitations": []
                    if selected and canonical_path is not None
                    else ["canonical mapping is incomplete"],
                }
            )
    return sorted(
        records, key=lambda row: (str(row["instrumentId"]), str(row["timeframe"]))
    )


def audit_volume_provenance_records(
    records: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    audited: list[dict[str, Any]] = []
    for source in records:
        row = dict(source)
        verification = verify_volume_semantics(row)
        audited.append({**row, "volumeSemantics": verification})
    counts = Counter(
        str(row["volumeSemantics"].get("route") or "E") for row in audited
    )
    payload: dict[str, Any] = {
        "schemaVersion": "volume_provenance_audit_v1",
        "datasetCount": len(audited),
        "verifiedExactTurnoverCount": counts["A"],
        "verifiedConservativeLowerBoundCount": counts["B"],
        "verifiedContractTurnoverCount": counts["C"],
        "independentProfileCount": counts["D"],
        "unavailableCount": counts["E"],
        "records": audited,
        "unitInferenceFromNumericDistribution": False,
    }
    payload["auditHash"] = stable_hash(payload, prefix="volume_provenance_audit")
    return payload


def build_exchange_identity_audit(
    *,
    research_exchange: str,
    ohlcv_exchange: str,
    funding_exchange: str,
    demo_execution_exchange: str,
) -> dict[str, Any]:
    identities = {
        "researchExchange": str(research_exchange),
        "ohlcvExchange": str(ohlcv_exchange),
        "fundingExchange": str(funding_exchange),
        "demoExecutionExchange": str(demo_execution_exchange),
    }
    same_exchange = (
        len(set(identities.values())) == 1
        and "unknown" not in identities.values()
        and "unverified_local_exchange" not in identities.values()
    )
    payload: dict[str, Any] = {
        "schemaVersion": "exchange_identity_and_portability_audit_v1",
        **identities,
        "sameExchange": same_exchange,
        "crossExchangePortabilityStatus": "verified" if same_exchange else "not_verified",
        "signalTimestampParity": "not_tested" if not same_exchange else "not_required",
        "returnCorrelation": "not_tested" if not same_exchange else "not_required",
        "eventOverlap": "not_tested" if not same_exchange else "not_required",
        "instrumentIdentity": "not_verified" if not same_exchange else "verified",
        "fundingDifference": "not_tested" if not same_exchange else "not_required",
        "releaseEligible": same_exchange,
        "knownLimitations": []
        if same_exchange
        else ["Cross-exchange portability has not been established."],
    }
    payload["auditHash"] = stable_hash(payload, prefix="exchange_identity_audit")
    return payload
